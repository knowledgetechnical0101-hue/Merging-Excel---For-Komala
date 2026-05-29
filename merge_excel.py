
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
import tempfile
import os

st.set_page_config(page_title="Advanced Excel Merger", layout="wide")

st.title("Advanced Excel File Merger")

st.write("""
Upload multiple Excel files having similar headers.
The program will:
- Automatically detect header row
- Merge all row data
- Handle blank cells
- Preserve formatting
- Download merged Excel file
""")

uploaded_files = st.file_uploader(
    "Upload Excel Files",
    type=["xlsx", "xlsm", "xls"],
    accept_multiple_files=True
)

# ---------------- HEADER DETECTOR ---------------- #

def detect_header_row(df):

    best_row = 0
    max_non_empty = 0

    for i in range(min(20, len(df))):

        row = df.iloc[i]

        non_empty_count = row.notna().sum()

        text_count = sum(
            isinstance(x, str) and str(x).strip() != ""
            for x in row
        )

        score = non_empty_count + text_count

        if score > max_non_empty:
            max_non_empty = score
            best_row = i

    return best_row


# ---------------- MAIN PROCESS ---------------- #

if uploaded_files:

    merged_df = pd.DataFrame()
    master_columns = None
    template_styles = {}

    try:

        for index, file in enumerate(uploaded_files):

            # Read raw excel
            raw_df = pd.read_excel(file, header=None)

            # Detect header row
            header_row = detect_header_row(raw_df)

            # Read actual data
            df = pd.read_excel(file, header=header_row)

            # Remove fully empty rows
            df = df.dropna(how="all")

            # Remove empty columns
            df = df.dropna(axis=1, how="all")

            # Store master headers
            if master_columns is None:
                master_columns = list(df.columns)

            # Match columns automatically
            df.columns = master_columns[:len(df.columns)]

            # Append all rows
            merged_df = pd.concat([merged_df, df], ignore_index=True)

            # Save formatting from first file
            if index == 0:

                temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
                temp_file.write(file.getvalue())
                temp_file.close()

                wb = load_workbook(temp_file.name)
                ws = wb.active

                for row in ws.iter_rows():

                    for cell in row:

                        if cell.has_style:

                            template_styles[cell.coordinate] = {
                                "font": cell.font.copy(),
                                "fill": cell.fill.copy(),
                                "border": cell.border.copy(),
                                "alignment": cell.alignment.copy()
                            }

                os.unlink(temp_file.name)

        # ---------------- CREATE OUTPUT ---------------- #

        output = BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:

            merged_df.to_excel(writer, index=False, sheet_name="Merged_Data")

            workbook = writer.book
            worksheet = writer.sheets["Merged_Data"]

            # Apply template formatting
            for row in worksheet.iter_rows():

                for cell in row:

                    ref = cell.coordinate

                    if ref in template_styles:

                        style = template_styles[ref]

                        cell.font = style["font"]
                        cell.fill = style["fill"]
                        cell.border = style["border"]
                        cell.alignment = style["alignment"]

            # Auto-adjust column width
            for column_cells in worksheet.columns:

                length = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in column_cells
                )

                worksheet.column_dimensions[
                    get_column_letter(column_cells[0].column)
                ].width = length + 5

        output.seek(0)

        st.success(f"{len(uploaded_files)} files merged successfully!")

        st.subheader("Merged Data Preview")
        st.dataframe(merged_df)

        st.download_button(
            label="Download Merged Excel File",
            data=output,
            file_name="Merged_Output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"Error: {e}")

else:
    st.info("Please upload Excel files.")

