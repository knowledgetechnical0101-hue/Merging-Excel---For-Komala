# app.py
import streamlit as st
import pdfplumber
import pandas as pd
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="SAP Invoice OCR Extractor", layout="wide")

st.title("📄 SAP Invoice OCR Extractor")
st.markdown("Upload SAP Invoice PDF files to extract invoice details and export to Excel.")

uploaded_files = st.file_uploader(
    "Upload SAP Invoice PDFs",
    type=["pdf"],
    accept_multiple_files=True
)

# ---------------------------------------------------------
# EXTRACT TEXT FROM PDF
# ---------------------------------------------------------

def extract_text_from_pdf(pdf_file):
    full_text = ""

    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            text = page.extract_text()

            if text:
                full_text += text + "\n"

    return full_text


# ---------------------------------------------------------
# EXTRACT INVOICE DETAILS
# ---------------------------------------------------------

def extract_invoice_details(text):

    invoice_data = {}

    # SAP No
    sap_match = re.search(r'SAP Entry no\.\s*(\d+)', text, re.IGNORECASE)
    invoice_data["SAP No"] = sap_match.group(1) if sap_match else ""

    # Invoice No
    invoice_match = re.search(r'(\d{14,})', text)
    invoice_data["Invoice No"] = invoice_match.group(1) if invoice_match else ""

    # Invoice Date
    date_match = re.search(r'Date\s+(\d{2}-[A-Za-z]{3}-\d{2})', text)
    invoice_data["Invoice Date"] = date_match.group(1) if date_match else ""

    # Round Off
    round_match = re.search(
        r'ZRND\s+Rounding Difference\s+(-?\d+\.\d+)',
        text
    )
    invoice_data["Round Off"] = (
        round_match.group(1) if round_match else "0"
    )

    # Grand Total
    grand_total_match = re.search(
        r'Total\s+([\d,]+\.\d+)',
        text
    )

    invoice_data["Grand Total"] = (
        grand_total_match.group(1).replace(",", "")
        if grand_total_match else ""
    )

    lines = text.splitlines()

    extracted_rows = []

    current_material = {}

    for i, line in enumerate(lines):

        # -------------------------------------------------
        # MATERIAL MAIN LINE
        # -------------------------------------------------

        material_pattern = re.search(
            r'^(\d+)\s+(\d+)\s+([A-Z0-9\-\s]+?)\s+(\d+\.\d+)\s+KL',
            line.strip()
        )

        if material_pattern:

            current_material = {}

            current_material["SAP No"] = invoice_data["SAP No"]
            current_material["Invoice No"] = invoice_data["Invoice No"]
            current_material["Invoice Date"] = invoice_data["Invoice Date"]

            current_material["Material Code"] = material_pattern.group(2)

            current_material["Description"] = (
                material_pattern.group(3).strip()
            )

            current_material["Quantity"] = (
                material_pattern.group(4)
            )

            # -------------------------------------------------
            # RATE
            # -------------------------------------------------

            rate_match = re.search(
                r'BASIC DESTINATION PRICE\s+\d+\.\d+\s+KL\s+([\d,]+\.\d+)',
                "\n".join(lines[i:i+5])
            )

            current_material["Rate"] = (
                rate_match.group(1).replace(",", "")
                if rate_match else ""
            )

            # -------------------------------------------------
            # LOCAL SALES TAX %
            # -------------------------------------------------

            lst_match = re.search(
                r'ZLST\s+Local sales tax\s+([\d\.]+)\s+\%',
                "\n".join(lines[i:i+8])
            )

            current_material["Local Sales Tax %"] = (
                lst_match.group(1)
                if lst_match else ""
            )

            # -------------------------------------------------
            # LOCAL SALES TAX VALUE
            # -------------------------------------------------

            lst_value_match = re.search(
                r'ZLST\s+Local sales tax\s+[\d\.]+\s+\%\s+([\d,]+\.\d+)',
                "\n".join(lines[i:i+8])
            )

            current_material["Local Sales Tax Value"] = (
                lst_value_match.group(1).replace(",", "")
                if lst_value_match else ""
            )

            # -------------------------------------------------
            # TOTAL FOR MATERIAL
            # -------------------------------------------------

            total_material_match = re.search(
                r'Total for material\s+([\d,]+\.\d+)',
                "\n".join(lines[i:i+10])
            )

            current_material["Total Value"] = (
                total_material_match.group(1).replace(",", "")
                if total_material_match else ""
            )

            current_material["Round Off"] = invoice_data["Round Off"]

            current_material["Grand Total"] = invoice_data["Grand Total"]

            extracted_rows.append(current_material)

    return extracted_rows


# ---------------------------------------------------------
# CREATE PROFESSIONAL EXCEL
# ---------------------------------------------------------

def create_excel(df):

    wb = Workbook()
    ws = wb.active
    ws.title = "SAP Invoice Report"

    # -----------------------------------------------------
    # TITLE
    # -----------------------------------------------------

    ws.merge_cells("A1:L1")

    title_cell = ws["A1"]

    title_cell.value = "SAP INVOICE OCR EXTRACTION REPORT"

    title_cell.font = Font(
        bold=True,
        size=16,
        color="FFFFFF"
    )

    title_cell.fill = PatternFill(
        "solid",
        fgColor="1F4E78"
    )

    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    # -----------------------------------------------------
    # ADD DATAFRAME
    # -----------------------------------------------------

    for row_idx, row in enumerate(
        dataframe_to_rows(df, index=False, header=True),
        start=3
    ):

        for col_idx, value in enumerate(row, start=1):

            cell = ws.cell(
                row=row_idx,
                column=col_idx,
                value=value
            )

            # HEADER STYLE
            if row_idx == 3:

                cell.font = Font(
                    bold=True,
                    color="FFFFFF"
                )

                cell.fill = PatternFill(
                    "solid",
                    fgColor="4F81BD"
                )

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            else:

                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center"
                )

            # BORDERS
            thin = Side(
                border_style="thin",
                color="000000"
            )

            cell.border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin
            )

    # -----------------------------------------------------
    # AUTO WIDTH
    # -----------------------------------------------------

    for col_idx, column_cells in enumerate(ws.columns, start=1):

        max_length = 0

        column = get_column_letter(col_idx)

        for cell in column_cells:

            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = max_length + 5

        ws.column_dimensions[column].width = adjusted_width

    # -----------------------------------------------------
    # SAVE EXCEL
    # -----------------------------------------------------

    excel_buffer = BytesIO()

    wb.save(excel_buffer)

    excel_buffer.seek(0)

    return excel_buffer


# ---------------------------------------------------------
# MAIN
# ---------------------------------------------------------

if uploaded_files:

    all_invoice_data = []

    with st.spinner("Processing SAP Invoices..."):

        for uploaded_file in uploaded_files:

            text = extract_text_from_pdf(uploaded_file)

            extracted_data = extract_invoice_details(text)

            if extracted_data:
                all_invoice_data.extend(extracted_data)

    if all_invoice_data:

        df = pd.DataFrame(all_invoice_data)

        st.success("✅ Invoice Data Extracted Successfully")

        st.dataframe(df, use_container_width=True)

        excel_file = create_excel(df)

        st.download_button(
            label="📥 Download Professional Excel Report",
            data=excel_file,
            file_name="SAP_Invoice_Extraction_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    else:

        st.error("❌ No invoice data detected.")


# ---------------------------------------------------------
# FOOTER
# ---------------------------------------------------------

st.markdown("---")
st.caption("Developed for SAP OCR Invoice Extraction")